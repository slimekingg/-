import random
import numpy as np
import openpyxl

def random_jidian():
    jidian = np.random.normal(3.0, 0.5, 90)
    for i in range(90):
        while jidian[i]>4:jidian[i]=np.random.normal(3.0, 0.5, 1)
        if jidian[i]<0:jidian[i]=0

    jidian.sort()
    return jidian

def random_attendance():
    
    taoke = []
    taokeren = random.randint(7,8)

    muqian = 0
    for i in range(90):
        if random.random()>0.5 :
            taoke.append(i)
            muqian+=1
        if muqian==taokeren : break

    cell = [[1 for _ in range(20)] for _ in range(90)]



    for i in range(20):
        suijitao = random.randint(0,3)
        suiji = []
        for j in range(suijitao):
            suiji.append(random.randint(0,89))

        for j in range(90):
            if suiji.count(j)>0 :
                cell[j][i]=0

    for i in range(90):
        if taoke.count(i)>0:
            suiji = []
            for j in range(2,20):suiji.append(j)
            suijishan=random.sample(suiji,4)
            for j in range(20):
                if suijishan.count(j):cell[i][j]=1
                else : cell[i][j]=0

    
    return cell 

def random_number():
    a = random.sample(range(32001101,32002650),90)
    for i in range(90):
        a[i]=str(a[i])
        a[i]=a[i].zfill(9)
    return a


for ii in range(1,6):
    cell = random_attendance()
    a = random_number()
    jidian = random_jidian()
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("课程"+str(ii),0)
    b = ["学号","出勤1","出勤2","出勤3","出勤4","出勤5","出勤6","出勤7","出勤8","出勤9","出勤10","出勤11","出勤12","出勤13","出勤14","出勤15","出勤16","出勤17","出勤18","出勤19","出勤20","绩点"]
    for i in range(90):ws.cell(i+2,22,jidian[i])
    for i in range(22):ws.cell(1,i+1,b[i])
    for i in range(90):ws.cell(i+2,1,a[i])
    for i in range(20):
        for j in range(90):
            ws.cell(j+2,i+2,cell[j][i])
    wb.save('点名情况课程'+str(ii)+'.xlsx')
